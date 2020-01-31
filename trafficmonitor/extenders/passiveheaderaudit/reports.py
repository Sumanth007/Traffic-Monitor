from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

font = Font(bold=True, name='Calibri')

cell_alignment = Alignment(
    horizontal='center', vertical='bottom', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0
)

cell_border = Border(
    left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'),
    top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'),
)

yellow_fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')


class PassiveHeaderAuditReport:
    def __init__(self, data, path):
        self.book = Workbook()
        self.path = path
        self.data = data
        self.generate_report()

    def generate_report(self):
        workbook_name = str(Path(f"{self.path}/Passive Secure Header Audit.xlsx"))

        headers = ['Date Time', 'Host', 'URL', 'Method', 'Status code', 'Response Header', 'Vulnerabilities']

        sheet = self.book.active
        sheet.title = "Results"
        sheet.append(headers)

        for indv_data in self.data:
            # print(indv_data)
            date_time = indv_data[0]
            host = indv_data[1]
            url = str(indv_data[2]).split("||")[0]
            method = indv_data[3]
            status_code = indv_data[4]
            rsp_header = indv_data[5]
            vulnerabilities = indv_data[6]
            sheet.append((date_time, host, url, method, status_code, rsp_header, vulnerabilities,))

        for cell in sheet[1:1]:
            cell.alignment = cell_alignment
            cell.border = cell_border
            cell.font = font
            cell.fill = yellow_fill

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    cell.alignment = cell_alignment
                    cell.border = cell_border

                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)

                    if cell.column is 'F' or cell.column is 'G':
                        cell.alignment = Alignment(horizontal='center', wrap_text=True)
                except:
                    pass

            adjust_width = (max_length + 2) * 1
            if column is 'F':
                sheet.column_dimensions[column].width = 70
            elif column is 'G':
                sheet.column_dimensions[column].width = 30
            else:
                sheet.column_dimensions[column].width = adjust_width

        self.book.save(workbook_name)
